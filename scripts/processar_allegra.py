"""
Processa o Excel Obra_Residencial Allegra.xlsx e gera o arquivo
no formato de importação do sistema de orçamento civil.

Correções aplicadas:
- Tipos de insumos (Produto→M, MO→MO, Equipamento→E)
- Codes para MO sem código
- Merge de composições de concreto manual (4 variantes → 1)
- Merge de composições de concreto usinado (2 variantes → 1)
- Nomes simplificados (elemento removido quando o sistema associa)
"""

import sys
import re
import uuid
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

SRC = r"C:\Users\PC\Downloads\Obra_Residencial Allegra.xlsx"
DST = r"C:\Users\PC\Downloads\Allegra_Importacao.xlsx"

# ---------------------------------------------------------------------------
# 1. Leitura do Excel fonte
# ---------------------------------------------------------------------------
print("Lendo arquivo fonte...")
df_comp_raw = pd.read_excel(SRC, sheet_name='COMPOSIÇÕES', header=0)
df_comp_raw.columns = [
    'CLASSIFICACAO','COD_AUX','COD_CP','DESC_COMP','DESC_INSUMO',
    'UM_INSUMO','COEF','QTD_INSUMO','CUSTO_MAT','CUSTO_SERV',
    'VAL_UN_MAT','VAL_UN_SERV','TIPO_ITEM','CODIGO_ITEM',
    'VALOR_TOTAL','TIPO','ETAPA'
]

df_ins_raw = pd.read_excel(SRC, sheet_name='INSUMOS', header=0)
df_ins_raw.columns = ['CODIGO','DESC','UM','VALOR','TIPO_ITEM']

# ---------------------------------------------------------------------------
# 2. Processar INSUMOS
# ---------------------------------------------------------------------------
print("Processando insumos...")

def inferir_tipo(codigo, descricao, tipo_original):
    d = str(descricao).strip().lower()
    tp = str(tipo_original).strip().lower()
    # Mão de obra
    if 'mão de obra' in d or 'm.o.' in d or d.startswith('mo ') or d.startswith('m.o'):
        return 'MO'
    if str(codigo).strip().lower().startswith('m') and str(codigo).strip() not in ('','nan'):
        # código M30, M8, M62 etc → MO
        try:
            int(str(codigo).strip()[1:])
            return 'MO'
        except:
            pass
    # Equipamentos
    if any(x in d for x in ['retroescavadeira','bomba concreto','perfuração rotativa','container']):
        return 'E'
    # Serviços (fornecimento+instalação completo)
    if 'fornecimento e instala' in d or 'forncecimento e instala' in d:
        return 'S'
    # Default: Material
    return 'M'

def inferir_categoria(descricao, tipo):
    d = str(descricao).lower()
    if tipo == 'MO':
        return 'Mão de Obra'
    if tipo == 'E':
        return 'Equipamento'
    if any(x in d for x in ['areia','brita','saibro']):
        return 'Areia e Brita'
    if any(x in d for x in ['cimento','cal hidratada','argamassa']):
        return 'Argamassa e Cimento'
    if any(x in d for x in ['vergalhão','arame','tela aço','treliça','espaçador']):
        return 'Aço e Ferragem'
    if any(x in d for x in ['madeira','tábua','tabua','escora','sarrafo']):
        return 'Madeira'
    if any(x in d for x in ['prego','parafuso','bucha']):
        return 'Material Básico'
    if any(x in d for x in ['impermeab','manta','vedacit','quartzolit']):
        return 'Impermeabilização'
    if any(x in d for x in ['pvc','tubo','cano','joelho','tê ','luva','adaptador','registro','válvula','caixa d','sifonada','gordura','eletroduto']):
        return 'Tubulação e Hidráulica'
    if any(x in d for x in ['cabo','disjuntor','interruptor','tomada','luminária','haste','quadro']):
        return 'Elétrico'
    if any(x in d for x in ['tijolo','bloco','canaleta']):
        return 'Alvenaria e Bloco'
    if any(x in d for x in ['telha','cumeeira','calha','rufo','funilaria']):
        return 'Cobertura e Telha'
    if any(x in d for x in ['porta','janela','esquadria','vidro','persiana','guarda']):
        return 'Esquadria'
    if any(x in d for x in ['tinta','selador','massa','lixa','pincel']):
        return 'Pintura e Verniz'
    if any(x in d for x in ['piso','porcelanato','cerâmico','granito','laminado','manta polietileno','rodapé']):
        return 'Revestimento e Piso'
    if any(x in d for x in ['drywall','perfil f5','chapa de dry','fita mesh','placomix','parafuso dry','tabica']):
        return 'Forros'
    if any(x in d for x in ['laje','concreto usinado','bomba concreto']):
        return 'Material Básico'
    return 'Material Básico'

insumos = []
mo_seq = 1

for _, row in df_ins_raw.iterrows():
    codigo = str(row['CODIGO']).strip() if pd.notna(row['CODIGO']) else ''
    desc = str(row['DESC']).strip() if pd.notna(row['DESC']) else ''
    um = str(row['UM']).strip() if pd.notna(row['UM']) else 'un'
    try:
        valor = float(str(row['VALOR']).strip()) if pd.notna(row['VALOR']) and str(row['VALOR']).strip() not in ('', ' ', 'nan') else 0.0
    except (ValueError, TypeError):
        valor = 0.0
    tipo_orig = str(row['TIPO_ITEM']).strip() if pd.notna(row['TIPO_ITEM']) else ''

    if not desc or desc.lower() in ('nan', 'insumos', ''):
        continue

    tipo = inferir_tipo(codigo, desc, tipo_orig)
    categoria = inferir_categoria(desc, tipo)

    # Atribuir código MO sequencial se não tiver
    if not codigo or codigo.lower() in ('nan', '0', ''):
        if tipo == 'MO':
            codigo = f'MO-{mo_seq:03d}'
            mo_seq += 1
        else:
            # insumo sem código → gerar M-
            codigo = f'M-AUTO-{uuid.uuid4().hex[:6].upper()}'

    insumos.append({
        'ID': str(uuid.uuid4()),
        'Código': codigo,
        'Descrição': desc,
        'Unidade': um if um not in ('nan', '', '0') else 'un',
        'Preço (R$)': round(valor, 2),
        'Tipo': tipo,
        'Categoria': categoria,
        'Status': 'ativo',
    })

# Remover duplicatas por código (manter primeiro)
seen_cod = {}
insumos_dedup = []
for ins in insumos:
    k = ins['Código'].strip().upper()
    if k not in seen_cod:
        seen_cod[k] = ins
        insumos_dedup.append(ins)

insumos = insumos_dedup
insumo_by_codigo = {i['Código'].strip(): i for i in insumos}
insumo_by_desc = {}
for i in insumos:
    insumo_by_desc[i['Descrição'].strip().lower()] = i

print(f"  → {len(insumos)} insumos processados")

# ---------------------------------------------------------------------------
# 3. Processar COMPOSIÇÕES
# ---------------------------------------------------------------------------
print("Processando composições...")

# Filtrar apenas linhas de composição real (desc_comp não nula e não "0")
df_items = df_comp_raw[
    df_comp_raw['DESC_COMP'].notna() &
    (df_comp_raw['DESC_COMP'].astype(str).str.strip() != '') &
    (df_comp_raw['DESC_COMP'].astype(str).str.strip() != '0') &
    (df_comp_raw['DESC_COMP'].astype(str).str.strip() != 'nan')
].copy()

# Obter composições únicas (linhas onde TIPO == 'Composição' ou é a primeira ocorrência do CP)
df_comp_headers = df_comp_raw[
    df_comp_raw['TIPO'].astype(str).str.strip() == 'Composição'
].copy()

# Normalização de nomes: correções de composições
CONSOLIDATIONS = {
    # Concreto manual FCK25 → um só (usar traço 6 sacos como padrão)
    'Concreto Manual em Obra FCK 25 Vigas Baldrame': 'Concreto Manual FCK 25',
    'Concreto Manual em Obra FCK 25 Blocos Sobre Estaca': 'Concreto Manual FCK 25',
    'Concreto Manual em Obra FCK 25 - Pilar ': 'Concreto Manual FCK 25',
    'Concreto Manual em Obra FCK 25 - Pilar': 'Concreto Manual FCK 25',
    'Concreto Manual em Obra FCK 25 - Escada ': 'Concreto Manual FCK 25',
    'Concreto Manual em Obra FCK 25 - Escada': 'Concreto Manual FCK 25',
    # Concreto usinado FCK25 → um só
    'Concreto Usinado FCK 25 - Entrepiso': 'Concreto Usinado FCK 25',
    'Concreto Usinado FCK 25 - Cobertura': 'Concreto Usinado FCK 25',
    # Formas - remover redundância de "Entrepiso/Cobertura" quando aplicável
    # (mantemos as que têm diferença estrutural real)
}

# Mapeamento etapa excel → etapa sistema
def etapa_sistema(etapa_raw):
    try:
        n = int(float(str(etapa_raw)))
        return f'{n:02d}'
    except:
        return '01'

# Coletar composições únicas
composicoes = {}  # desc_normalizada → {id, codigo, descricao, etapa, ...}

for _, row in df_comp_headers.iterrows():
    desc_orig = str(row['DESC_COMP']).strip()
    if not desc_orig or desc_orig in ('0', 'nan'):
        continue
    cod_cp = str(row['COD_CP']).strip() if pd.notna(row['COD_CP']) else ''
    etapa = etapa_sistema(row['ETAPA'])

    # Aplicar consolidação de nome
    desc = CONSOLIDATIONS.get(desc_orig, desc_orig).strip()

    if desc not in composicoes:
        composicoes[desc] = {
            'ID': str(uuid.uuid4()),
            'Código': cod_cp.replace('.', '_') if cod_cp not in ('nan', '') else f'CP-{uuid.uuid4().hex[:6].upper()}',
            'Descrição': desc,
            'Unidade Produção': 'm²',  # padrão; será ajustado abaixo
            'Produção': 1,
            'Descrição Técnica': '',
            'Status': 'ativo',
            'etapa': etapa,
        }

# Ajustar unidade produção por tipo de composição
UNIDADE_PROD = {
    'concreto': 'm³',
    'armadura': 'kg',
    'forma': 'm²',
    'laje': 'm²',
    'alvenaria': 'm²',
    'chapisco': 'm²',
    'emboço': 'm²',
    'reboco': 'm²',
    'revestimento': 'm²',
    'impermeabiliz': 'm²',
    'pintura': 'm²',
    'contrapiso': 'm²',
    'cobertura': 'm²',
    'telhamento': 'm²',
    'trama': 'm²',
    'forro': 'm²',
    'estaca': 'un',
    'bloco sobre': 'un',
    'escada': 'un',
    'piso': 'm²',
    'rodapé': 'm',
    'reaterro': 'm³',
    'movimentação': 'm³',
    'limpeza': 'm²',
    'placa': 'un',
    'tapume': 'm²',
    'depósito': 'vb',
    'instalação provisória': 'vb',
    'poste': 'un',
    'enfiação': 'm',
    'eletroduto': 'm',
    'tubos': 'm',
    'conexões': 'vb',
    'caixas': 'vb',
    'interruptor': 'un',
    'tomada': 'un',
    'luminária': 'un',
    'quadro': 'un',
    'louças': 'vb',
    'metais': 'vb',
    'locação': 'm²',
    'verga': 'm',
    'pingadeira': 'm',
    'soleira': 'm',
    'granito': 'm²',
    'guarda-corpo': 'm²',
    'vidro': 'm²',
    'esquadria': 'm²',
    'massa': 'm²',
    'calhas': 'm',
    'reservatório': 'un',
    'limpeza final': 'vb',
}

for desc, c in composicoes.items():
    d = desc.lower()
    for k, u in UNIDADE_PROD.items():
        if k in d:
            c['Unidade Produção'] = u
            break

print(f"  → {len(composicoes)} composições únicas")

# ---------------------------------------------------------------------------
# 4. Processar ITENS de composição
# ---------------------------------------------------------------------------
print("Processando itens das composições...")

# Linhas de itens: onde há DESC_INSUMO preenchida e TIPO != 'Composição'
df_it = df_items[
    df_items['DESC_INSUMO'].notna() &
    (df_items['DESC_INSUMO'].astype(str).str.strip() != '') &
    (df_items['DESC_INSUMO'].astype(str).str.strip() != 'nan') &
    (df_items['TIPO'].astype(str).str.strip() != 'Composição')
].copy()

itens_comp = []
seen_comp_insumo = set()  # para dedup dentro de composição consolidada

for _, row in df_it.iterrows():
    desc_comp_orig = str(row['DESC_COMP']).strip()
    desc_comp = CONSOLIDATIONS.get(desc_comp_orig, desc_comp_orig).strip()

    if desc_comp not in composicoes:
        continue

    comp_id = composicoes[desc_comp]['ID']

    desc_insumo = str(row['DESC_INSUMO']).strip()
    codigo_item = str(row['CODIGO_ITEM']).strip() if pd.notna(row['CODIGO_ITEM']) else ''

    # Tentar achar o insumo pelo código
    ins = None
    if codigo_item and codigo_item not in ('nan', '0', ''):
        ins = insumo_by_codigo.get(codigo_item.strip())
    if ins is None:
        ins = insumo_by_desc.get(desc_insumo.lower())

    if ins is None:
        # Tentar match parcial para MO sem código
        for key, val in insumo_by_desc.items():
            if desc_insumo.lower() in key or key in desc_insumo.lower():
                ins = val
                break

    if ins is None:
        continue

    insumo_id = ins['ID']

    # Chave de dedup: evitar duplicar item consolidado
    dedup_key = (comp_id, insumo_id)
    if dedup_key in seen_comp_insumo:
        continue
    seen_comp_insumo.add(dedup_key)

    try:
        coef = float(str(row['COEF']).strip()) if pd.notna(row['COEF']) and str(row['COEF']).strip() not in ('', ' ', 'nan') else 0
    except (ValueError, TypeError):
        coef = 0
    um = str(row['UM_INSUMO']).strip() if pd.notna(row['UM_INSUMO']) else ins['Unidade']

    # Para Concreto Manual FCK 25 consolidado: usar coef do primeiro (baldrame/escada = 6 sacos padrão)
    # O pilar tinha 4 sacos — normalizamos para 6 (traço FCK25 padrão)
    if desc_comp == 'Concreto Manual FCK 25' and 'cimento' in desc_insumo.lower():
        coef = 6.0  # 6 sacos/m³ FCK25

    # Para Concreto Usinado FCK 25 consolidado: usar coef padrão 0.083
    if desc_comp == 'Concreto Usinado FCK 25' and codigo_item in ('P32', 'P26'):
        coef = round(1/12, 4)

    itens_comp.append({
        'ID': str(uuid.uuid4()),
        'ID Composição': comp_id,
        'ID Insumo': insumo_id,
        'Coeficiente': round(coef, 6),
        'Unidade': um if um not in ('nan', '', '0') else ins['Unidade'],
    })

print(f"  → {len(itens_comp)} itens de composição")

# ---------------------------------------------------------------------------
# 5. Gerar Excel de importação
# ---------------------------------------------------------------------------
print("Gerando Excel de importação...")

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT = Font(name='Arial', size=9)
ALT_FILL = PatternFill('solid', start_color='EEF3F7')

def write_sheet(wb, name, columns, rows, first=False):
    if first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 20

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            if fill:
                cell.fill = fill

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    return ws

# Sheet INSUMOS
ins_cols = ['ID', 'Código', 'Descrição', 'Unidade', 'Preço (R$)', 'Tipo', 'Categoria', 'Status']
write_sheet(wb, 'INSUMOS', ins_cols, insumos, first=True)

# Sheet Composições
comp_rows = [{
    'ID': c['ID'],
    'Código': c['Código'],
    'Descrição': c['Descrição'],
    'Unidade Produção': c['Unidade Produção'],
    'Produção': c['Produção'],
    'Descrição Técnica': c['Descrição Técnica'],
    'Status': c['Status'],
} for c in composicoes.values()]
comp_cols = ['ID', 'Código', 'Descrição', 'Unidade Produção', 'Produção', 'Descrição Técnica', 'Status']
write_sheet(wb, 'Composições', comp_cols, comp_rows)

# Sheet Itens_Composição
itens_cols = ['ID', 'ID Composição', 'ID Insumo', 'Coeficiente', 'Unidade']
write_sheet(wb, 'Itens_Composição', itens_cols, itens_comp)

wb.save(DST)
print(f"\n✓ Arquivo gerado: {DST}")
print(f"  - {len(insumos)} insumos")
print(f"  - {len(composicoes)} composições")
print(f"  - {len(itens_comp)} itens de composição")

# Resumo das consolidações
print("\n=== COMPOSIÇÕES CONSOLIDADAS ===")
merged = {}
for desc_orig, desc_new in CONSOLIDATIONS.items():
    merged.setdefault(desc_new, []).append(desc_orig)
for new, olds in merged.items():
    print(f"  '{new}'  ←  {len(olds)} composições:")
    for o in olds:
        print(f"       - {o}")
